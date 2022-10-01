from flask import Flask,render_template,request,redirect,flash
import phone_serach
import os
from werkzeug.utils import secure_filename
import time

app = Flask(__name__)
app.config['UPLOAD_FOLDER'] = 'uploads'
app.config['SECRET_KEY'] = 'thee7tyeiutrandomi7retqeitugkebftring'   

@app.route('/',methods=["GET","POST"])
def index():

    data = False
    if request.method == "POST":
        phone = request.form.get("search_data")
        data,res = phone_serach.main(phone)
        if data.get("isError"):
            flash(data.get("error").get("message"))
            data,res = False

    return render_template('single.html',data=data)


ALLOWED_EXTENSIONS = ['xlsx']
def allowed_file(filename):
    return '.' in filename and \
           filename.rsplit('.', 1)[1].lower() in ALLOWED_EXTENSIONS


@app.route('/upload',methods=["GET","POST"])
def upload_file():
    d_count = 1
    emp_list = []
    if request.method == 'POST':

        if 'file' not in request.files:
            flash('No file part')
            return redirect(request.url)
        file = request.files['file']

        if file.filename == '':
            flash('No selected file')
            return redirect(request.url)
        if file and allowed_file(file.filename):
            filename = secure_filename(file.filename)
            file.save(os.path.join(app.config['UPLOAD_FOLDER'], filename))
            import openpyxl
            dataframe = openpyxl.load_workbook(os.path.abspath(os.path.join(app.config['UPLOAD_FOLDER'], filename)))
            dataframe1 = dataframe.active
            for row in range(0, dataframe1.max_row):
                for col in dataframe1.iter_cols(1, dataframe1.max_column):
                    if col[row].value:

                        if col[row].value.lower() == "phone number":
                            pass
                        else:
                            d_count += 1
                            store = col[row].value
                            data,res = phone_serach.main(store)
                            if res == 200:
                                emp_list.append(data)

                            if d_count == 30:
                                d_count = 0
                                print("Sleeping....")

                                time.sleep(20)
            print(emp_list)
            return render_template("multiple.html",dummy_data=emp_list)
            # return emp_list
    return render_template("multiple.html")



if __name__ == "__main__":
    app.run()


